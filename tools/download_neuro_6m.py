import os
import sys
import re
import json
import time
from pathlib import Path
from datetime import datetime, date
from typing import Optional, Callable

# Add Dcom to JPG directory to sys.path
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from dcom_pipeline import (
    HOSPITALS,
    search_patient_studies,
    download_studies_list,
    run_pipeline,
    resolve_study_viewer_url,
    _safe_name,
    _study_date_token,
    _normalise_dicom_date,
    CLINICAL,
)

sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)
sys.stderr.reconfigure(encoding='utf-8', line_buffering=True)

BASE_NEURO_DIR = Path(r"D:\Học tập\NỘI TRÚ THẦN KINH\Google drive\Working\NeuroOncology\BN mổ NeuroOncology")
EXCEL_FILE = BASE_NEURO_DIR / "Danh sách BN Phẫu thuật U NÃO và U TỦY 6 Tháng 2026.xlsx"
CHECKPOINT_FILE = PROJECT_ROOT / "neuro_6m_checkpoint.json"

def sanitize_filename(name: str) -> str:
    """Sanitize string for Windows filename."""
    if not name:
        return ""
    cleaned = re.sub(r'[<>:"/\\|?*]', '_', str(name))
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned

def parse_date_flexible(date_val) -> Optional[date]:
    """Parse date from various Excel/RIS string/datetime formats."""
    if not date_val:
        return None
    if isinstance(date_val, (datetime, date)):
        return date_val.date() if isinstance(date_val, datetime) else date_val
    s = str(date_val).strip()
    
    # Format: YYYY-MM-DD or YYYY.MM.DD or YYYY/MM/DD
    m_iso = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m_iso:
        y, m, d = int(m_iso.group(1)), int(m_iso.group(2)), int(m_iso.group(3))
        try:
            return date(y, m, d)
        except ValueError:
            pass

    # Format: DD-MM-YYYY or DD.MM.YYYY or DD/MM/YYYY
    m_vn = re.search(r"(\d{1,2})[-./](\d{1,2})[-./](\d{4})", s)
    if m_vn:
        d, m, y = int(m_vn.group(1)), int(m_vn.group(2)), int(m_vn.group(3))
        try:
            return date(y, m, d)
        except ValueError:
            pass

    # Format: YYYYMMDD
    m_num = re.search(r"\b(\d{4})(\d{2})(\d{2})\b", s)
    if m_num:
        y, m, d = int(m_num.group(1)), int(m_num.group(2)), int(m_num.group(3))
        try:
            return date(y, m, d)
        except ValueError:
            pass
    return None

def determine_surgery_stage(study_date: Optional[date], surg_date: Optional[date]) -> str:
    """Classify the imaging study relative to surgery date."""
    if not study_date or not surg_date:
        return ""
    delta = (study_date - surg_date).days
    if delta < 0:
        if delta < -45:
            return "trước mổ (khám sớm)"
        return "trước mổ"
    elif delta in (0, 1):
        return "sau mổ 24h"
    elif delta == 2:
        return "sau mổ 48h"
    elif 3 <= delta <= 10:
        return f"sau mổ {delta} ngày"
    elif 20 <= delta <= 45:
        return "sau mổ 1 tháng"
    elif 46 <= delta <= 75:
        return "sau mổ 2 tháng"
    elif 76 <= delta <= 110:
        return "sau mổ 3 tháng"
    elif 111 <= delta <= 200:
        return "sau mổ 6 tháng"
    elif delta > 200:
        return "sau mổ tái khám"
    return f"sau mổ {delta} ngày"

def normalize_modality(raw_mod: str, desc: str = "") -> str:
    """Normalize modality to standard MRI or CLVT."""
    raw = (raw_mod or "").strip().upper()
    d = (desc or "").strip().upper()
    if raw in ("MR", "MRI") or "MR" in raw or "CỘNG HƯỞNG TỪ" in d or "CONG HUONG TU" in d:
        return "MRI"
    if raw in ("CT", "CLVT", "CAT") or "CT" in raw or "CẮT LỚP" in d or "CAT LOP" in d:
        return "CLVT"
    if raw in ("DX", "CR", "RG", "XQ", "X-RAY") or "XQUANG" in d or "X-QUANG" in d:
        return "XQuang"
    return raw or "UNKNOWN"

def load_excel_patients():
    """Load and parse the 97 patients from the 6-month NeuroOncology Excel."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Tất cả U NÃO & U TỦY (119 ca)"]
    headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
    
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        pid = str(d.get("Mã NB") or "").strip()
        if not pid or pid == "None":
            continue
        pname = str(d.get("Họ và tên BN") or "").strip()
        age = str(d.get("Tuổi") or "").strip()
        sex = str(d.get("Giới tính") or "").strip()
        surg_time = str(d.get("Th/g bắt đầu mổ") or "").strip()
        surg_date = parse_date_flexible(surg_time)
        diag = str(d.get("Chẩn đoán sau PT") or "").strip()
        proc = str(d.get("Phương pháp phẫu thuật") or "").strip()
        mbh = str(d.get("Kết quả Mô bệnh học 1") or "").strip()
        month = str(d.get("Tháng Phẫu thuật") or "").strip()

        patients.append({
            "pid": pid,
            "name": pname,
            "age": age,
            "sex": sex,
            "surg_time": surg_time,
            "surg_date": surg_date,
            "diag": diag,
            "proc": proc,
            "mbh": mbh,
            "month": month,
        })
    return patients

def find_existing_patient_folder(pid: str) -> Optional[Path]:
    """Find existing folder for a patient ID in BASE_NEURO_DIR."""
    for item in BASE_NEURO_DIR.iterdir():
        if item.is_dir() and item.name not in ("Data EMR", "scratch"):
            m = re.match(r"^(\d{8,10})", item.name)
            if m and m.group(1) == pid:
                return item
            if item.name.startswith(f"{pid}-") or item.name.startswith(f"{pid} ") or item.name.startswith(f"{pid}_"):
                return item
    return None

def build_patient_folder_name(p: dict) -> str:
    """Build standard patient folder name: Mã BN - Họ tên - Tuổi - Chẩn đoán."""
    pid = p['pid']
    name = sanitize_filename(p['name'])
    age_str = f"{p['age']}T" if p.get('age') else ""
    diag_str = sanitize_filename(p.get('diag', ''))[:45]
    
    parts = [pid, name]
    if age_str:
        parts.append(age_str)
    if diag_str:
        parts.append(diag_str)
    return "-".join(parts)

def analyze_existing_subfolders(patient_folder: Path, surg_date: Optional[date], dhy_studies: list[dict]):
    """Analyze subfolders in patient folder to detect DHY studies vs outside hospital studies."""
    if not patient_folder.exists():
        return []

    dhy_study_dates = set()
    for s in dhy_studies:
        s_date = parse_date_flexible(s.get("date"))
        if s_date:
            dhy_study_dates.add(s_date)

    existing_subs = []
    for sub in patient_folder.iterdir():
        if not sub.is_dir() or sub.name.startswith("."):
            continue
        if sub.name.upper() in ("CONTENT", "DICOM", "DATA EMR", "REPORTS", "ATTACHMENTS"):
            continue

        sub_date = parse_date_flexible(sub.name)
        files = list(sub.rglob("*"))
        jpg_count = sum(1 for f in files if f.is_file() and f.suffix.lower() in (".jpg", ".jpeg", ".png"))
        dcm_count = sum(1 for f in files if f.is_file() and (f.suffix.lower() == ".dcm" or "." not in f.name))
        
        is_on_dhy = sub_date in dhy_study_dates if sub_date else False

        sub_upper = sub.name.upper()
        if "CLVT" in sub_upper or "CT" in sub_upper:
            mod = "CLVT"
        elif "MRI" in sub_upper or "MR" in sub_upper:
            mod = "MRI"
        else:
            mod = "UNKNOWN"

        stage = determine_surgery_stage(sub_date, surg_date)

        existing_subs.append({
            "folder": sub,
            "name": sub.name,
            "date": sub_date,
            "modality": mod,
            "stage": stage,
            "jpg_count": jpg_count,
            "dcm_count": dcm_count,
            "is_on_dhy": is_on_dhy,
            "is_outside": not is_on_dhy,
        })
    return existing_subs

def generate_patient_index(patient_folder: Path, p_info: dict, all_studies_meta: list[dict]):
    """Write or update unified patient-index.json."""
    index_path = patient_folder / "patient-index.json"
    now_iso = datetime.now().astimezone().isoformat()
    
    studies_dict = {}
    for s in all_studies_meta:
        uid = s.get("studyUid") or s.get("study_uid") or f"study_{s.get('date')}_{s.get('modality')}"
        studies_dict[uid] = s

    manifest = {
        "format": "dcom-patient-index-v1",
        "patientId": p_info.get("pid", ""),
        "patientName": p_info.get("name", ""),
        "patientAge": p_info.get("age", ""),
        "patientSex": p_info.get("sex", ""),
        "diagnosis": p_info.get("diag", ""),
        "surgeryDate": p_info.get("surg_time", ""),
        "hospitalKey": "dhy",
        "hospitalName": "BV Đại học Y Hà Nội",
        "updatedAt": now_iso,
        "studies": studies_dict
    }
    
    if not index_path.exists():
        manifest["createdAt"] = now_iso
    else:
        try:
            old = json.loads(index_path.read_text(encoding="utf-8"))
            manifest["createdAt"] = old.get("createdAt", now_iso)
        except Exception:
            manifest["createdAt"] = now_iso

    index_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest

def process_single_patient(p: dict, log=print, dry_run=False):
    """Process a single patient: search DHY RIS, preserve outside studies, download missing DHY studies."""
    pid = p["pid"]
    pname = p["name"]
    surg_date = p["surg_date"]
    
    log(f"\n{'='*70}")
    log(f"▶ XỬ LÝ BỆNH NHÂN: {pid} - {pname} (Tuổi: {p.get('age')}, Mổ: {p.get('surg_time')})")
    log(f"{'='*70}")

    # 1. Determine or locate patient folder
    existing_folder = find_existing_patient_folder(pid)
    if existing_folder:
        patient_folder = existing_folder
        log(f"✓ Sử dụng thư mục hiện có: {patient_folder.name}")
    else:
        folder_name = build_patient_folder_name(p)
        patient_folder = BASE_NEURO_DIR / folder_name
        if not dry_run:
            patient_folder.mkdir(parents=True, exist_ok=True)
        log(f"✓ Tạo thư mục mới: {patient_folder.name}")

    # 2. Search DHY RIS for studies
    log(f"🔎 Đang tìm kiếm các ca chụp của BN {pid} trên RIS BV Đại học Y Hà Nội...")
    try:
        dhy_studies = search_patient_studies(
            hospital_key="dhy",
            patient_id=pid,
            modality="ALL",
            log=log,
            headless=True
        )
    except Exception as e:
        log(f"❌ Lỗi khi tìm kiếm trên RIS ĐHY: {e}")
        dhy_studies = []

    log(f"-> Tìm thấy {len(dhy_studies)} ca chụp trên RIS ĐHY.")

    # Filter only relevant MR and CT studies
    filtered_dhy_studies = []
    for s in dhy_studies:
        m = (s.get("modality") or "").upper()
        desc = (s.get("desc") or "").upper()
        if m in ("MR", "MRI", "CT", "CLVT") or "MR" in m or "CT" in m or "CONG HUONG TU" in desc or "CAT LOP" in desc:
            filtered_dhy_studies.append(s)

    log(f"-> Số ca MRI / CT cần xử lý: {len(filtered_dhy_studies)}")

    # 3. Analyze existing subfolders
    existing_subs = analyze_existing_subfolders(patient_folder, surg_date, filtered_dhy_studies)
    log(f"📁 Thư mục hiện có {len(existing_subs)} ca chụp cục bộ:")
    for sub_info in existing_subs:
        tag = "🇻🇳 [VIỆN NGOÀI / BẢO TỒN]" if sub_info["is_outside"] else "🏥 [ĐẠI HỌC Y]"
        log(f"   • {sub_info['name']} | JPGs: {sub_info['jpg_count']} | {tag}")

    all_studies_meta = []

    # Record preserved outside studies
    for sub_info in existing_subs:
        if sub_info["is_outside"]:
            log(f"   🛡️ BẢO TỒN NGUYÊN VẸN phim viện ngoài: {sub_info['name']}")
            all_studies_meta.append({
                "studyUid": f"outside_{sub_info['name']}",
                "date": str(sub_info['date']) if sub_info['date'] else "",
                "modality": sub_info['modality'],
                "description": sub_info['name'],
                "folder": sub_info['name'],
                "stage": sub_info['stage'],
                "source": "outside_hospital",
                "status": "complete",
                "imageCount": sub_info['jpg_count'],
                "preservedAt": datetime.now().astimezone().isoformat()
            })

    # 4. Download / Update DHY Studies
    for idx, st in enumerate(filtered_dhy_studies, 1):
        st_date_str = st.get("date", "")
        st_date = parse_date_flexible(st_date_str)
        st_mod = normalize_modality(st.get("modality", ""), st.get("desc", ""))
        st_desc = st.get("desc", "")
        stage = determine_surgery_stage(st_date, surg_date)
        
        # Build desired subfolder name: YYYY-MM-DD - MRI/CLVT - stage (or desc)
        date_prefix = str(st_date) if st_date else "KHONG_RO_NGAY"
        if stage:
            desired_sub_name = f"{date_prefix} - {st_mod} - {stage}"
        else:
            clean_desc = sanitize_filename(st_desc)[:35]
            desired_sub_name = f"{date_prefix} - {st_mod} - {clean_desc}" if clean_desc else f"{date_prefix} - {st_mod}"

        # Check if already downloaded in existing subfolder with same date
        matched_existing = None
        for sub_info in existing_subs:
            if not sub_info["is_outside"] and sub_info["date"] == st_date:
                matched_existing = sub_info
                break

        target_dir = patient_folder / desired_sub_name
        
        if matched_existing and matched_existing["jpg_count"] > 50:
            log(f"   ✓ Ca {idx} ({st_date_str} {st_mod}): Đã có sẵn đầy đủ ảnh ({matched_existing['jpg_count']} JPGs) tại '{matched_existing['name']}'.")
            target_sub_folder = matched_existing["folder"]
            if matched_existing["name"] != desired_sub_name and not target_dir.exists():
                try:
                    matched_existing["folder"].rename(target_dir)
                    target_sub_folder = target_dir
                    log(f"     -> Chuẩn hóa tên subfolder thành: {desired_sub_name}")
                except Exception as e:
                    log(f"     -> Giữ tên folder: {matched_existing['name']} ({e})")
            
            all_studies_meta.append({
                "studyUid": st.get("study_uid") or st.get("uid"),
                "date": st_date_str,
                "modality": st_mod,
                "description": st_desc,
                "folder": target_sub_folder.name,
                "stage": stage,
                "source": "dhy_ris",
                "status": "complete",
                "imageCount": matched_existing["jpg_count"],
                "updatedAt": datetime.now().astimezone().isoformat()
            })
            continue

        if dry_run:
            log(f"   [DRY RUN] Sẽ tải ca {idx}: {st_date_str} {st_mod} ({st_desc}) -> {desired_sub_name}")
            continue

        # Download study from DHY RIS
        log(f"\n   ⬇️ ĐANG TẢI CA {idx}/{len(filtered_dhy_studies)}: {st_date_str} {st_mod} ({st_desc})")
        log(f"      Giai đoạn mổ xác định: '{stage}'")
        log(f"      Thư mục lưu: {desired_sub_name}")

        try:
            viewer_url = resolve_study_viewer_url(
                hospital_key="dhy",
                study_uid=st.get("study_uid") or st.get("uid"),
                log=log,
                headless=True
            )
            target_dir.mkdir(parents=True, exist_ok=True)
            
            dl, cv, jpg_dir = run_pipeline(
                url=viewer_url,
                out_base=target_dir,
                log=log,
                headless=True,
                quality=100,
                save_png=False,
                contrast_mode=CLINICAL,
                rename_patient_root=False,
                jpg_folder_name_override="JPG"
            )
            
            img_count = cv.converted if hasattr(cv, "converted") else (cv.total() if hasattr(cv, "total") else (int(cv) if isinstance(cv, (int, float)) else 0))
            log(f"   ✓ Tải và chuyển đổi thành công: {dl.total() if hasattr(dl, 'total') else dl} DICOM, {img_count} JPG.")
            all_studies_meta.append({
                "studyUid": st.get("study_uid") or st.get("uid"),
                "date": st_date_str,
                "modality": st_mod,
                "description": st_desc,
                "folder": desired_sub_name,
                "stage": stage,
                "source": "dhy_ris",
                "status": "complete",
                "imageCount": img_count,
                "downloadedAt": datetime.now().astimezone().isoformat()
            })
        except Exception as e:
            log(f"   ❌ Lỗi khi tải ca {idx} ({st.get('study_uid')}): {e}")
            all_studies_meta.append({
                "studyUid": st.get("study_uid") or st.get("uid"),
                "date": st_date_str,
                "modality": st_mod,
                "description": st_desc,
                "folder": desired_sub_name,
                "stage": stage,
                "source": "dhy_ris",
                "status": f"error: {e}",
                "imageCount": 0
            })

    # 5. Generate / Update patient-index.json
    if not dry_run:
        manifest = generate_patient_index(patient_folder, p, all_studies_meta)
        log(f"✓ Đã ghi file patient-index.json với {len(manifest.get('studies', {}))} ca chụp.")

    return True

def load_checkpoint() -> dict:
    """Load progress checkpoint from JSON file."""
    if CHECKPOINT_FILE.exists():
        try:
            return json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def save_checkpoint(data: dict) -> None:
    """Save progress checkpoint to JSON file."""
    try:
        CHECKPOINT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"⚠ Lỗi lưu checkpoint: {e}")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Download and harmonize NeuroOncology 6 months 2026 data.")
    parser.add_argument("--dry-run", action="store_true", help="Perform dry run scan only without downloading.")
    parser.add_argument("--start", type=int, default=1, help="1-based start patient index.")
    parser.add_argument("--limit", type=int, default=None, help="Number of patients to process.")
    parser.add_argument("--pid", type=str, default=None, help="Process a single specific patient ID.")
    parser.add_argument("--force", action="store_true", help="Force re-processing even if marked completed in checkpoint.")
    args = parser.parse_args()

    patients = load_excel_patients()
    print(f"Tổng số bệnh nhân trong file Excel 6 tháng: {len(patients)}")

    checkpoint = load_checkpoint()

    if args.pid:
        patients = [p for p in patients if p['pid'] == args.pid]
        print(f"Lọc theo mã BN {args.pid}: {len(patients)} bệnh nhân.")
    else:
        start_idx = max(0, args.start - 1)
        end_idx = start_idx + args.limit if args.limit is not None else len(patients)
        patients = patients[start_idx:end_idx]
        print(f"Phạm vi xử lý: từ BN số {args.start} đến {start_idx + len(patients)} (Tổng: {len(patients)} BN)")

    completed_count = sum(1 for p in patients if checkpoint.get(p['pid'], {}).get("status") == "completed")
    print(f"Đã hoàn thành từ trước: {completed_count}/{len(patients)} BN.")

    for idx, p in enumerate(patients, 1):
        pid = p['pid']
        if not args.force and not args.dry_run and checkpoint.get(pid, {}).get("status") == "completed":
            print(f"[{idx}/{len(patients)}] ⏭️ BỎ QUA BN {pid} - {p['name']} (Đã hoàn thành trước đó)")
            continue

        print(f"\n[{idx}/{len(patients)}] Đang xử lý: {pid} - {p['name']}")
        try:
            success = process_single_patient(p, log=print, dry_run=args.dry_run)
            if success and not args.dry_run:
                checkpoint[pid] = {
                    "status": "completed",
                    "name": p['name'],
                    "age": p.get('age'),
                    "diag": p.get('diag'),
                    "surg_time": p.get('surg_time'),
                    "completedAt": datetime.now().astimezone().isoformat()
                }
                save_checkpoint(checkpoint)
        except Exception as exc:
            print(f"❌ LỖI NGHIÊM TRỌNG KHI XỬ LÝ BN {pid}: {exc}")
            import traceback
            traceback.print_exc()

    print("\n" + "="*70)
    print("HOÀN TẤT QUÁ TRÌNH XỬ LÝ.")
    print("="*70)
    print("="*70)

if __name__ == "__main__":
    main()
